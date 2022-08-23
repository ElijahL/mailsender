from mailsender import MailSender
import traceback

MAIL_ADDRESS="elia.zmushko@gmail.com"
MAIL_PASSWORD="lbrwqxxziofaxxcf"
SMTP_SERVER="smtp.gmail.com"
PORT=587

EMAIL_COLUMN_NAME = "Email"
NAME_COLUMN_NAME  = "First Name"

def send_mails(xl_path: str, cc: str, subject: str, body: str, attachments: str):
    try:
        sender = MailSender(
            username=MAIL_ADDRESS,
            password=MAIL_PASSWORD,
            smtp_server=SMTP_SERVER,
            port=PORT
            )

        from openpyxl import load_workbook
        wb = load_workbook(xl_path)
        ws = wb.active

        column_access = {}
        for idx, cell in enumerate(ws[1]):
            column_access[cell.value] = idx
            
        for i in range(2, ws.max_row + 1):
            row = ws[i]
            if ws.row_dimensions[row[0].row].hidden == False:
                sender.send_mail(
                    sender=MAIL_ADDRESS,
                    receiver=row[column_access[EMAIL_COLUMN_NAME]].value,
                    cc=cc,
                    subject=subject,
                    body=body.format(name=row[column_access[NAME_COLUMN_NAME]].value),
                    attachment_paths=attachments
                    )
                print(f'Sent to {row[column_access[EMAIL_COLUMN_NAME]].value}')

    except Exception:
        traceback.print_exc()
        return False
    else:
        return True
